"""
This program sends gene symbols to the PubMed database to retrieve the number of citations.

7/7/16
UPDATES:
- separates keywords by AND in one search
- added option to sort or not sort
- added option to limit the number of genes to search, using radiobuttons
- added functionality of Option button in Advanced menu to lead you back to the main menu if clicked again
- added advanced section in config.ini file
- made warning that the config file was outdated if new sections or options were added after a previously
saved version
- updated reset button in the edit menu
"""

import threading
import time
import os
from urllib.error import URLError

from tkinter import *
import tkinter.ttk as ttk
import tkinter.filedialog, tkinter.messagebox
from Bio import Entrez
from openpyxl import load_workbook
from openpyxl.utils import _get_column_letter
from openpyxl.comments import Comment
import requests
import mygene

import layout

wb = None
filename = None
save_as_name = None
pb_int = 0  # global variable for num of times get_count has been executed
total_queries = 0
total_count_col = 29
symbol_col = 'G'
ask_quit = False
num_genes = 0

# TODO make word cloud based below (show descriptions in word cloud)
# TOdO we want low ratio (one decimal) high count


def progress(root):
    root.bar.pb.grid()
    root.bar.start()


def get_entries(root):
    global wb, total_queries, total_count_col, num_genes

    if filename is None:
        tkinter.messagebox.showinfo(title='Error', message="Please select a file to sort.")
        return
    email = root.custom_frames['FormPage'].ents['Email'].get()
    keywords = root.custom_frames['FormPage'].ents['Keywords'].get()

    keywords = [x.strip() for x in keywords.split(',')] if keywords != '' else []
    # the input is turned into a list of keywords
    wb = try_file(filename)
    ws = wb.active
    num_genes = int(root.custom_frames['AdvancedPage'].entry.get()) if root.custom_frames['AdvancedPage'].v.get() == \
        "SELECT" or root.custom_frames['AdvancedPage'].entry.get() > ws.max_row - 1 else ws.max_row - 1
    rows, total_count_col = read_sheet(ws, num_genes)
    try:
        genes = get_aliases(rows[1:], root)
        print("Num genes: %d" % num_genes)
        total_queries = num_genes * 2 if len(keywords) > 0 else num_genes
        print("Total queries: %d" % total_queries)
        if root.custom_frames['AdvancedPage'].desc.get() == 1:  # if the box is checked
            total_queries += num_genes  # adds the number of comment queries
        if root.custom_frames['AdvancedPage'].v.get() == "SELECT":
            print("Adding 20 genes to output")
            ws = wb.create_sheet(title='Output', index=0)
            r = 1
            for row in rows:
                c = 1
                for data in row:
                    ws.cell(row=r, column=c).value = data
                    c += 1
                r += 1
        try:
            root.custom_frames['FormPage'].b3.unbind('<Button-1>')
            root.custom_frames['FormPage'].b3.unbind('<Enter>')
            root.custom_frames['FormPage'].b3.config(background='dark khaki')
            thread1 = threading.Thread(target=lambda: progress(root))
            thread2 = threading.Thread(target=lambda: set_info(ws, email, keywords, genes, root))
            thread1.start()  # TODO look into this error (google app engine)
            thread2.start()
        except Exception as e:
            tkinter.messagebox.showerror(title='Error',
                                         message='The process was interrupted, but your file was saved.\n' + str(e))
            if ".xlsx" == save_as_name[-5:]:
                save_as = save_as_name
            else:
                save_as = save_as_name + ".xlsx"
            wb.save(save_as)
            sys.exit()

    except (AttributeError, IndexError) as e:
        tkinter.messagebox.showinfo(title="Column Error",
                                    message="Check advanced page column input. \n" + str(e))


def try_file(user_input):
    """Loops until user enters an xlsx file (with or without the extension) that exists in the specified directory and
    returns the workbook in the file."""
    user_input = str(user_input)
    if ".xlsx" == user_input[-5:]:
        file_name = user_input
    else:
        file_name = user_input + '.xlsx'
    return load_workbook(filename=file_name, data_only=True)


def read_sheet(ws, amt=None):
    """Reads all existing values from list of columns if amt is not specified, otherwise, it gets up to the amt
    parameter"""
    if amt is None:
        amt = ws.max_row - 1
    total_col = ws.max_column + 1
    gene_rows = []
    for row in range(1, amt+2):  # amt+2 selects two blank rows
        aliases = []
        for col in range(1, total_col):
            cell = ws.cell(row=row, column=col).value
            aliases.append(cell)
        gene_rows.append(aliases)
    return gene_rows, total_col


def get_aliases(gene_rows, root):
    """This converts the data from the rows into all of the gene's aliases. First it adds the symbol to the list of
    aliases, then it adds the synonyms (separated by semicolon) and returns all of the aliases of all the genes."""
    global symbol_col
    symbol_input = root.custom_frames["AdvancedPage"].ents["Symbol Column"].get()
    synonym_input = root.custom_frames["AdvancedPage"].ents["Synonyms Column"].get()
    symbol_col = symbol_input if symbol_input != "" else 'G'
    synonym_col = synonym_input if synonym_input != "" else 'H'
    symbol_col = ord(symbol_col.lower()) - 97  # convert the column letter to a number
    synonym_col = ord(synonym_col.lower()) - 97
    gene_aliases = []
    for gene in gene_rows:
        aliases = [gene[symbol_col]]  # symbols column is added to the list of names
        if gene[synonym_col] is not None:
            synonyms = gene[synonym_col].split("; ")
            aliases.extend(synonyms)
        gene_aliases.append(aliases)
    return gene_aliases


def set_info(ws, email, keywords, genes, root):
    """Converts each gene to a Entrez ID then to its full gene name from the my gene library and passes them to the
    get_count method. Then each count is entered into the worksheet. If a gene is not found in the database, it is
    appended to the list of genes not found, which is later printed out and sorted to the top of the worksheet."""
    global pb_int
    checkbox = root.custom_frames["AdvancedPage"].desc.get()
    quick_save = False
    all_counts = []
    number = 1
    for aliases in genes:  # for each list of aliases (one gene) in the full list of genes
        if ask_quit:
            print("Quitting...")
            sys.exit()
        print("#%d" % number)
        counts = get_count(aliases, keywords, email, root)
        # the length of the list counts returned is the length of keywords + 1
        if len(counts) < 2 if len(keywords) > 1 else 1:  # if counts list is not complete
            quick_save = True
            break
        else:
            print(counts)
            all_counts.append(counts)
            number += 1

    write_info(ws, all_counts, keywords)

    if not quick_save:

        if len(keywords) > 0:  # set the ratio column
            ws.column_dimensions[_get_column_letter(total_count_col+2)].width = 16
            col = total_count_col + 2
            ws.cell(row=1, column=col).value = "COUNT RATIO"
            row = 2
            for counts in all_counts:
                try:
                    count = int(counts[1]) / int(counts[0])
                except ZeroDivisionError:
                    count = 1
                ws.cell(row=row, column=col).value = count
                row += 1
        if root.custom_frames["AdvancedPage"].sort.get() == 1:
            ws, rows = sort_ws(ws)
        else:
            rows = read_sheet(ws)

    if checkbox == 1:
        print("Getting descriptions...")
        column1 = [row[symbol_col] for row in rows]
        row = 2
        for symbol in column1:
            if ask_quit:
                sys.exit()
            pb_int += 1
            if symbol != '':
                try:
                    comment = Comment(get_summary(symbol), "PubMed")
                    comment.width = '200pt'
                    comment.height = '100pt'
                    ws.cell(row=row, column=symbol_col+1).comment = comment
                except Exception as e:
                    tkinter.messagebox.showerror(title='Error',
                                                 message='Getting descriptions was interrupted by an error, ' +
                                                         'but your spreadsheet was saved.')
            row += 1

    wb.save(save_as_name)

    print("Done!")
    if tkinter.messagebox.showinfo(title='Success',
                                   message="Your file is located in " + os.path.dirname(filename)) == 'ok':
        root.bar.pb.grid_forget()


def get_count(aliases, keywords, email, root):
    """Returns the number of times the gene comes up on PubMed. In the first try loop, the code decides whether it is
    getting the total count or narrow count. If it is the second time and the total count was 0, then the narrow count
    is automatically 0 so it will lessen the number of queries sent to PubMed. The try statement's purpose is to except
    a TypeError in case there is a mistake in the list (i.e. a date). Then the query is sent to PubMed and it gets the
    "Count" line."""
    global pb_int
    Entrez.email = email
    counts = []  # first number is total, second number is narrowed by keyword, third number is next keyword, etc
    for i in range(2) if len(keywords) > 1 else range(1):
        pb_int += 1
        try:
            if i == 0:
                query = " OR ".join(aliases)
            elif counts[0] == "0":
                counts.append("0")
                continue  # if total count is 0, then narrow count is automatically 0
            else:
                query = "(%s) AND %s" % (" OR ".join(aliases), " AND ".join(keywords))
        except TypeError:  # if the gene name is in the wrong format (like if a date is entered accidentally)
            counts.append("0")
            continue
        try:
            handle = Entrez.egquery(term=query)
        except URLError as e:
            print(str(e))
            time.sleep(5)
            try:
                handle = Entrez.egquery(term=query)
            except Exception as e:
                tkinter.messagebox.showwarning(title="Error", message=str(e)+" Your partial output has been saved.")
                return counts
        print(query)
        record = Entrez.read(handle)
        for row in record["eGQueryResult"]:
            counts.append(row["Count"])
            break
    time.sleep(.5)
    return counts


def write_info(ws, all_counts, keywords):
    ws.column_dimensions[_get_column_letter(total_count_col)].width = 16
    ws.cell(row=1, column=total_count_col).value = "TOTAL COUNT"
    ws.column_dimensions[_get_column_letter(total_count_col+1)].width = 16
    ws.cell(row=1, column=total_count_col+1).value = '"%s" COUNT' % "/".join(keywords)  # to title the columns
    row = 2
    for counts in all_counts:
        col = total_count_col
        for count in counts:
            ws.cell(row=row, column=col).value = int(count)  # for each count loop it will move one row down
            col += 1
        row += 1


def sort_ws(ws):
    """Sorts the file's genes into a list of tuples"""
    rows, _ = read_sheet(ws)  # gets a list of genes from read_gene
    rows = rows[1:num_genes+1]
    try:
        rows.sort(key=lambda x: x[total_count_col-1])  # sorts genes by "Total Count" column
    except TypeError:
        print("Sorting TypeError")
        return ws, rows
    row = 2
    for gene in rows:
        col = 1
        for cell in gene:
            # writes sorted gene dictionary into the file to replace the existing values
            ws.cell(row=row, column=col).value = cell
            col += 1
        row += 1
    return ws, rows


def get_summary(symbol):
    version = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11'
    mg = mygene.MyGeneInfo()
    try:
        entrez_id = mg.query('symbol:%s' % symbol, species='human')['hits'][0]['entrezgene']
    except (KeyError, IndexError):
        return "No entries found. (Entrez ID not found)"
    url = 'http://www.ncbi.nlm.nih.gov/gene/' + str(entrez_id)

    response = requests.get(url, version)
    html_output = response.text

    search_string_start = '<dt>Summary</dt>'
    match_start = html_output.find(search_string_start)
    if match_start != -1:
        match_start += len(search_string_start)
        html_output = html_output[match_start:]

        search_string_end = '<dt>Orthologs</dt>'
        match_end = html_output.find(search_string_end)
        if match_end != -1:
            html_output = html_output[:match_end]
            extract_string = re.sub('<[^<]+?>', '', html_output)  # takes out the HTML tags
        else:
            extract_string = "No entries found. (match_end = -1)"

    else:
        extract_string = "No entries found. (match_start= -1)"
    return extract_string


def main():
    root = layout.Window('BioDataSorter')
    root.geometry('450x330+300+300')
    root.mainloop()

if __name__ == '__main__':
    main()
