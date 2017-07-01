#!/usr/bin/env python

# -*- coding: utf-8 -*-

"""
This program sends gene symbols to the PubMed database to retrieve the number
of citations.

This is a project that we are working on for Winthrop University Hospital for
genes to study in diabetes and multiple sclerosis. The program takes data in
the form of Excel spreadsheets that have a gene 'Symbol' column and 'Synonyms'
column, like the spreadsheets that can be downloaded from NCBI's Gene
Expression Omnibus (GEO), which is a public functional genomics data repository
 for array-based and sequence-based data.


Attributes:
    wb (openpyxl.Workbook): This variable is used to store the representation
        of the Excel file chosen chosen as input, which is changed, then stored
        under another name that the user selects.
        
    pb_int (int): Keeps track of the progress of the process based on the 
        number of queries so far, and updates progress bar accordingly.

    form_elements (dict):
        filename (str): The absolute file path and name of the input file
            selected by the user.
        email (str): The user's email to be sent with each Entrez query.
        save_as_name (str): The absolute file path and name of the output file
            selected or created by the user.
        column_letters ([str]): A list of strings (one letter long for column
            letter or "AUTO") provided by the user of the symbol and synonyms
            columns. Are initialized as None.
        num_genes  (int): The total number of genes to query. This is used to
            get the total number of queries. It is multiplied by two or three
            depending on the selection of the Descriptions box. This is also
            initialized as None.
        descriptions (bool): This is changed to True if the user checks the
            descriptions box, but is false otherwise.
        sort (bool): This is changed to True if the user checks the sort box,
            but is false otherwise.
            
    col_num (dict):
        symbol (int): index of symbol column in row
        synonym (int): index of synonyms column in row



"""

from datetime import datetime
from socket import timeout
from time import sleep, clock
from tkinter.messagebox import showwarning, showinfo, showerror
from os import path
from threading import Thread
from tkinter import *
from urllib import request
from urllib.error import URLError
import logging

from mygene import MyGeneInfo
from requests import get
from Bio import Entrez
from openpyxl import load_workbook
from openpyxl.comments import Comment

import layout

__version__ = "2.0.3"

# =======
# GLOBALS
# =======

logger = logging.getLogger()
wb = None
pb_int = 0  # global variable for num of times get_count has been executed
total_queries = 0
total_count_col = 29
ask_quit = False
ask_save_and_quit = False
start_time = None

# this is a global variable so the user input can be saved right after the
# program is run
form_elements = {
    'filename': None,  # includes path and name
    'email': None,
    'keywords': None,
    'save_as_name': None,
    'column_letters': None,
    'num_genes': None,
    'descriptions': None,
    'sort': None
}
col_num = {
    'symbol': None,
    'synonym': None
}


def get_entries(root):
    global wb, total_queries, total_count_col, form_elements, start_time

    def is_number(s):
        try:
            float(s)
            return True
        except ValueError:
            return False

    def internet_on():
        try:
            # tries pubmed.gov ip
            request.urlopen('http://130.14.29.109', timeout=1)
            return True
        except (URLError, timeout) as err:
            logging.info("Can't connect to wifi: " + str(err))
            return False

    if not internet_on():
        showwarning(title='Connect to internet',
                          message="Make sure you are connected to "
                          "internet before starting.")
        return

    form_page = root.custom_frames['FormPage']
    advanced_page = root.custom_frames['AdvancedPage']
    # set all values to global variable
    form_elements['email'] = form_page.ents['Email'].get()
    keywords_str = form_page.ents['Keywords'].get()
    form_elements['keywords'] = [x.strip() for x in keywords_str.split(',')] \
        if keywords_str != '' else []

    if form_elements['filename'] is None:
        showinfo(title='Error', message="Please select a file to sort.")
        return

    # the input is turned into a list of keywords
    try:
        wb = try_file(form_elements['filename'])
    except TypeError:
        showinfo(title='Error', message="Unexpected spreadsheet type.")
        return
    except FileNotFoundError as e:
        showinfo(title='Error', message=str(e))

    # Nothing is wrong so proceed

    # Start timer
    clock()

    # remove key if it is there
    if root.key.winfo_ismapped() == 1:
        root.key.pack_forget()
    ws = wb.active
    if advanced_page.v.get() == 'SELECT':

        if advanced_page.entry.get() == '' or not \
                is_number(advanced_page.entry.get()):
            showinfo(title='Invalid input',
                     message='Insert a number for \'Top x genes\'.')
            return
        if int(advanced_page.entry.get()) >= ws.max_row:
            form_elements['num_genes'] = ws.max_row - 1
        else:
            form_elements['num_genes'] = int(advanced_page.entry.get())
    else:
        form_elements['num_genes'] = ws.max_row - 1
    if advanced_page.auto_manual_columns.get() == 'AUTO':
        form_elements['column_letters'] = 'AUTO'
    else:
        form_elements['column_letters'] = [advanced_page.symbol_col.get(),
                                           advanced_page.synonyms_col.get()]
    form_elements['descriptions'] = advanced_page.desc.get() == 1
    form_elements['sort'] = advanced_page.sort.get() == 1

    # try:
    rows = read_sheet(ws)
    rows = remove_duplicates(rows)
    rows[0][col_num['symbol']] = 'Gene symbol'  # error if output is used
    rows[0][col_num['synonym']] = 'Gene title'

    if not rows:
        logging.info("quit get_entries b/c rows is None")
        return

    # cuts down to the number of genes that the user wants
    rows = rows[:form_elements['num_genes']+1]

    # this is the column that TOTAL COUNT will be written in
    total_count_col = len(rows[0])
    genes = get_aliases(rows)
    form_elements['num_genes'] = len(genes)
    logging.info("Num genes: %d" % form_elements['num_genes'])
    total_queries = form_elements['num_genes'] * 2
    if advanced_page.desc.get() == 1:  # if the box is checked

        # adds the number of comment queries
        total_queries += form_elements['num_genes']
    logging.info("Total queries: %d" % total_queries)
    ws = wb.create_sheet(title='Output', index=0)
    logging.info(col_num['symbol'])
    write_rows(rows, ws, col_num['symbol'])

    try:
        advanced_page.b3.config(state="disabled")
        form_page.run_button.config(state='disabled')
        thread1 = Thread(target=root.bar.start)
        thread2 = Thread(target=lambda: set_info(ws,
                                                 form_elements['email'],
                                                 form_elements['keywords'],
                                                 genes,
                                                 root))
        thread1.start()
        thread2.start()
    except Exception as e:
        showerror(title='Error',
                  message='The process was interrupted, but your file was '
                          'saved.\n' + str(e))
        if ".xlsx" == form_elements['save_as_name'][-5:]:
            save_as = form_elements['save_as_name']
        else:
            save_as = form_elements['save_as_name'] + ".xlsx"
        wb.save(save_as)
        sys.exit()


def try_file(user_input):
    """Loops until user enters an xlsx file (with or without the extension)
    that exists in the specified directory and returns the workbook in the
    file."""
    user_input = str(user_input)
    if ".xlsx" == user_input[-5:]:
        file_name = user_input
    else:
        file_name = user_input + '.xlsx'
    return load_workbook(filename=file_name, data_only=True)


def read_sheet(ws, amt=None):
    """Reads all existing values from list of columns if amt is not specified,
    otherwise, it gets up to the amt parameter"""
    if amt is None:
        amt = ws.max_row - 1
    gene_rows = []
    for row in range(1, amt+2):
        cells = []
        for col in range(1, ws.max_column+1):
            cell = ws.cell(row=row, column=col).value
            cells.append(cell)
        gene_rows.append(cells)
    return gene_rows


def locate_columns(rows):
    """Sets col_num global variable to correct symbol and synonyms indices.
    
    If automatic column retrieval is selected, this searches for a common 
    header of the symbol and synonym column. Otherwise, it searches for the
    header that the user manually inputted.
    
    :param rows: List of lists of each row of the Excel columns
    :return: void
    """
    global col_num

    def lower_list(l):
        return [i.lower() for i in l if i]

    # if the setting is auto for columns
    # searches rows ignoring capitalization
    if form_elements['column_letters'] == 'AUTO':
        if 'gene symbol' in lower_list(rows[0]):
            col_num['symbol'] = lower_list(rows[0]).index('gene symbol')
        elif 'symbol' in lower_list(rows[0]):
            col_num['symbol'] = lower_list(rows[0]).index('symbol')
        else:
            showinfo(title="Column Error",
                     message='Automatic column search could not find "Gene '
                             'symbol" or "SYMBOL" in spreadsheet')
            return

    # if the column letters setting is manual
    else:
        symbol_letter = form_elements['column_letters'][0]

        # ensures that the columns are valid for symbol and synonyms
        if ' ' in symbol_letter:
            showinfo(title='Column Error', message="Check advanced page symbol"
                                                   " column input for spaces.")
            return
        elif any(char.isdigit() for char in symbol_letter):
            showinfo(title='Column Error',
                     message="Check advanced page symbol column input for "
                             "numbers.")
            return

        # if every element in the column is None (empty)
        elif all(el is None for el in rows[ord(symbol_letter.lower()) - 97]):
            showinfo(title='Column Error',
                     message="Check that the advanced page symbol column input"
                             " was in the range of the spreadsheet.")
            return

        # no problems
        else:
            col_num['symbol'] = ord(symbol_letter.lower()) \
                             - 97

    if form_elements['column_letters'] == 'AUTO':
        if 'gene title' in lower_list(rows[0]):
            col_num['synonym'] = lower_list(rows[0]).index('gene title')
        elif 'SYNONYMS' in rows[0]:
            col_num['synonym'] = lower_list(rows[0]).index('SYNONYMS')
        else:
            showinfo(title='Column Error',
                     message='Automatic column search could not find '
                             '"Gene title" or "SYNONYMS" in spreadsheet')
    else:
        if ' ' in form_elements['column_letters'][1]:
            showinfo(title='Column Error', message="Check advanced page symbol"
                                                   " column input for spaces.")
            return
        elif any(char.isdigit() for char in
                 form_elements['column_letters'][1]):
            showinfo(title='Column Error',
                     message="Check advanced page symbol column input for "
                             "numbers.")
            return

        # if every element in the column is None (empty)
        elif all(el is None for el in
                 [row[ord(form_elements['column_letters'][1].lower()) - 97]
                  for row in rows]):
            showinfo(title='Column Error',
                     message="Check that the advanced page symbol column input"
                             " was in the range of the spreadsheet.")
            return
        else:
            col_num['synonym'] = ord(form_elements['column_letters'][1].lower())\
                               - 97


def remove_duplicates(rows):
    """Removes duplicated rows and empty rows from worksheet"""
    global form_elements
    rows_no_duplicates = []
    [rows_no_duplicates.append(row) for row in rows
        if row not in rows_no_duplicates]
    rows = [row for row in rows_no_duplicates if len([data for data in row if
                                                      data is not None]) > 1]

    locate_columns(rows)
    # and removes rows with no symbol
    rows = [row for row in rows if row[col_num['symbol']] is not None]

    return rows


def write_rows(rows, ws, symbol_col=None):
    """Writes rows into new worksheet and moves symbols column to the front.
    
    :param rows: List of rows in the spreadsheet to write onto the output sheet
    :param ws: openpyxl.Worksheet for output information
    :param symbol_col: int to move symbol column to front. Would be None for 
        conversion to Excel file from tab-delineated text file.
    :return: void
    """

    col = 1
    # if this test failed there would be a blank column
    if symbol_col is not None:
        r = 1
        # put symbols column first
        for row in rows:
            ws.cell(row=r, column=col).value = row.pop(symbol_col)
            r += 1
        col = 2  # start in the second column because first column is symbol

    r = 1
    # fill in rest of the Excel spreadsheet
    for row in rows:
        c = col
        for data in row:
            ws.cell(row=r, column=c).value = data
            c += 1
        r += 1


def get_aliases(gene_rows):
    """This converts the data from the rows into all of the gene's aliases.
    First it adds the symbol to the list of aliases, then it adds the synonyms
    (separated by semicolon) and returns all of the aliases of all the
    genes."""

    gene_aliases = []
    for gene in gene_rows[1:]:
        if type(gene[col_num['symbol']]) == datetime:
            aliases = []
        else:
            # symbols column is added to the list of names
            aliases = gene[col_num['symbol']].split('///')
        if gene[col_num['synonym']] is not None:

            # AttributeError: 'datetime.datetime' object has no attribute
            # 'split'
            synonyms = str(gene[col_num['synonym']]).split("; ")
            aliases.extend(synonyms)
        gene_aliases.append(aliases)
    return gene_aliases


def colnum_string(n):
    div = n
    string = ""
    while div > 0:
        m = (div - 1) % 26
        string = chr(65 + m) + string
        div = int((div - m) / 26)
    return string


def set_info(ws, email, keywords, genes, root):
    """Writes TOTAL COUNT column and %keyword% COLUMN (helper function
    _write_info), as well as descriptions and sorting if these options are
    selected by the user."""
    global pb_int, form_elements
    quick_save = False
    all_counts = []
    number = 1

    # for each list of aliases (one gene) in the full list of genes
    for aliases in genes:
        if ask_quit:
            logging.info("Quitting...")
            sys.exit()
        if ask_save_and_quit:
            quick_save = True
            logging.info("Saving...")
            break
        logging.info("#%d" % number)

        # makes sure no aliases are common words that throw off the search
        # (string length is longer than 2 letters)
        aliases = ["(" + alias + ")" for alias in aliases if len(alias) > 2]
        counts = get_count(aliases, keywords, email)
        # the length of the list counts returned is the length of keywords + 1
        if len(counts) < 2:  # if counts list is not complete
            quick_save = True
            break
        else:
            logging.info(counts)
            all_counts.append(counts)
            number += 1

    ws, rows = _write_info(ws, all_counts, keywords)

    if not quick_save:
        if form_elements['sort']:
            if root.custom_frames['AdvancedPage'].reverse_sort.get() == 1:
                ws_rows = sort_ws(rows, True)
            else:
                ws, rows = sort_ws(rows)

        # sets the ratio column
        col = total_count_col + 2
        ws.column_dimensions[colnum_string(col)].width = 16
        ws.cell(row=1, column=col).value = "COUNT RATIO"
        row = 2
        # TODO ratio does not work check test14!!!

        # sorts all_counts so it matches the sorted file
        all_counts = sorted(all_counts, key=lambda el: int(el[0]))
        for counts in all_counts:
            try:
                count = int(counts[1]) / int(counts[0])
            except ZeroDivisionError:
                count = 0  # divide by zero error always means 0/0
            ws.cell(row=row, column=col).value = count
            row += 1

        if form_elements['descriptions']:
            logging.info("Getting descriptions...")

            # symbol col is now 0
            symbols_list = [row[0] for row in rows[1:]]
            row = 2
            for i, symbol in enumerate(symbols_list):
                if ask_quit:
                    sys.exit()
                pb_int += 1
                if symbol != '':
                    try:
                        comment = Comment(get_summary(symbol), "PubMed")
                        comment.width = '500pt'  # TODO see if this works
                        comment.height = '700pt'  # original was 108 x 59.25

                        # assuming symbols column is in the first column
                        ws.cell(row=row,
                                column=1).comment = comment
                    except Exception as e:
                        error_msg = ("Getting descriptions was interrupted by"
                                     " an error, but your spreadsheet was "
                                     "saved.")
                        showerror(title='Error', message=error_msg)
                        logging.info(str(e))
                        break
                row += 1

    else:
        logging.info("Quick save")

    wb.save(form_elements['save_as_name'])
    total_time = clock()
    total_time_str = "Total time: " + str(int(total_time / 60)) + " min " + \
                     str(int(total_time % 60)) + " sec "

    logging.info("Done! " + total_time_str)
    if showinfo(title='Success',
                message=total_time_str +
                "Your file is located in " + path.dirname(
                    form_elements['save_as_name'])) == 'ok':
        root.bar.pb.pack_forget()
        root.custom_frames['FormPage'].reset()
        root.custom_frames['AdvancedPage'].reset()
        pb_int = 0


def get_count(aliases, keywords, email):
    """Returns the number of times the gene comes up on PubMed. In the first
    try loop, the code decides whether it is getting the total count or narrow
    count. If it is the second time and the total count was 0, then the narrow
    count is automatically 0 so it will lessen the number of queries sent to
    PubMed. The try statement's purpose is to except a TypeError in case there
    is a mistake in the list (i.e. a date). Then the query is sent to PubMed
    and it gets the "Count" line."""

    global pb_int
    Entrez.email = email
    counts = []  # first number is total, second number is narrowed by keywords
    for i in range(2):

        # increment the progressbar which will update as the other thread runs
        pb_int += 1
        try:
            if i == 0:
                query = " OR ".join(aliases)
            elif counts[0] == "0":
                counts.append("0")

                # if total count is 0, then narrow count is automatically 0
                continue
            else:
                query = "(%s) AND %s" % (" OR ".join(aliases),
                                         " AND ".join(keywords))

        # if the gene name is in the wrong format (like if a date)
        except TypeError:
            counts.append("0")
            continue

        try:
            handle = Entrez.egquery(term=query)
        except URLError as e:
            logging.info(str(e))
            # if PubMed blocks the queries then it waits 5 seconds and repeats
            sleep(5)
            try:
                handle = Entrez.egquery(term=query)
            except Exception as e:
                showwarning(title="Error",
                            message=str(e)+" Your partial output "
                            "has been saved.")
                return counts
        logging.info(query)
        record = Entrez.read(handle)
        for row in record["eGQueryResult"]:
            counts.append(row["Count"])
            break
    sleep(.5)
    return counts


def _write_info(ws, all_counts, keywords):
    ws.column_dimensions[colnum_string(total_count_col)].width = 16

    ws.cell(row=1, column=total_count_col).value = "TOTAL COUNT"
    ws.column_dimensions[colnum_string(total_count_col+1)].width = 16
    ws.cell(row=1, column=total_count_col+1).value = \
        '"%s" COUNT' % "/".join(keywords)  # to title the columns
    row = 2
    for counts in all_counts:
        col = total_count_col

        # for each count loop it will move one row down
        for count in counts:
            ws.cell(row=row, column=col).value = int(count)
            col += 1
        row += 1

    rows = read_sheet(ws)
    return ws, rows


def sort_ws(rows, sort_reverse=False):
    """Sorts the file's genes into a list of tuples"""
    ws = wb.active
    header = rows[0]
    gene_rows = rows[1:]

    # sorts genes by "Total Count" column
    gene_rows.sort(key=lambda x: x[total_count_col - 1], reverse=sort_reverse)
    row = 2
    for gene in gene_rows:
        col = 1
        for cell in gene:
            # writes sorted gene dictionary into the file to replace the
            # existing values
            ws.cell(row=row, column=col).value = cell
            col += 1
        row += 1
    rows = [header]
    rows.extend(gene_rows)
    return ws, rows


def get_summary(symbol):
    version = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) ' \
              'Gecko/20071127 Firefox/2.0.0.11'
    mg = MyGeneInfo()
    try:
        entrez_id = mg.query('symbol:%s' % symbol,
                             species='human')['hits'][0]['entrezgene']
    except Exception as e:
        logging.info("Error with query: " + str(e))
        return "No entries found. (Entrez ID not found)"

    url = 'http://www.ncbi.nlm.nih.gov/gene/' + str(entrez_id)
    response = get(url, version)
    html_output = response.text

    search_string_start = '<dt>Summary</dt>'
    match_start = html_output.find(search_string_start)
    if match_start != -1:
        match_start += len(search_string_start)
        html_output = html_output[match_start:]

        search_string_end = '<dt>Orthologs</dt>'
        match_end = html_output.find(search_string_end)
        if match_end != -1:
            html_output = html_output[:match_end]

            # takes out the HTML tags
            extract_string = re.sub('<[^<]+?>', '', html_output)
        else:
            extract_string = "No entries found. (match_end = -1)"

    else:
        extract_string = "No entries found. (match_start = -1)"
    return extract_string


def set_log():
    formatter = '%(asctime)s %(levelname)s %(message)s'
    logging.basicConfig(level=logging.DEBUG,
                        format=formatter,
                        filename='app.log',
                        filemode='w')
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(ch)


def main():
    set_log()
    root = layout.Window('BioDataSorter')
    root.geometry(str(layout.WINDOW_WIDTH) + 'x' + str(layout.WINDOW_HEIGHT)
                  + '+300+300')
    root.mainloop()

if __name__ == '__main__':
    main()
